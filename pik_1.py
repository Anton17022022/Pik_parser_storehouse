import requests
from bs4 import BeautifulSoup
import time
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from multiprocessing import Pool
from pathlib import Path
from shutil import rmtree
import openpyxl as ol
import json
import random

url = 'https://www.pik.ru/search/storehouse'

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0'
}

problem_content_pages = []


def check_consist_pages_def():

    check_consist_pages = []

    list = os.listdir(
        r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\data_pages')

    for item in list:
        check_consist_pages.append(
            f'https://www.pik.ru/search/{item[:-4]}/storehouse')

    if os.path.exists(r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\problem_content.txt') is True:
        os.remove(
            r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\problem_content.txt')

    with open(r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\pik_links.txt', 'r') as file:
        for item in file:

            if item[:-1] not in check_consist_pages:
                problem_content_pages.append(item[:-1])

    with open(r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\problem_content.txt', 'w') as file:
        for item in problem_content_pages:
            file.write(f'{item}\n')


def check_pages():
    new_list_of_links = []
    old_list_of_links = []
    with open(r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\pik_links.txt', 'r') as file:
        for item in file:
            new_list_of_links.append(item[:-1])

    with open(r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\old_list_pages.txt', 'r') as file:
        for item in file:
            old_list_of_links.append(item[:-1])

    if os.path.exists(r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\problem_pages.txt'):
        os.remove(
            r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\problem_pages.txt')
    if os.path.exists(r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\old_list_pages.txt'):
        os.remove(
            r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\old_list_pages.txt')

    for item in new_list_of_links:
        if item not in old_list_of_links:
            with open(r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\problem_pages.txt', 'w') as file:
                file.write(f'{item}\n')

    with open(r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\old_list_pages.txt', 'w') as file:
        for item in new_list_of_links:
            file.write(f'{item}\n')


def download_pages_objects(url):
    if os.path.isfile(r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\pik_links.txt') == True:
        os.remove(
            r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\pik_links.txt')

    for path in Path(r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\data_pages').glob('*'):
        if path.is_dir():
            rmtree(path)
        else:
            path.unlink()

    list_links = []
    req = requests.get(url, headers=headers)
    soup = BeautifulSoup(req.text, "html5lib")

    for i in soup.find_all("a", class_="styles__ProjectCard-uyo9w7-0 friPgx"):
        list_links.append('https://www.pik.ru'+i.get('href')+'\n')

    list_links = list(set(list_links))

    with open(r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\pik_links.txt', 'a') as file:
        for link in list_links:
            file.write(link)

    check_pages()


def get_list_objects_links(url):
    download_pages_objects(url)
    list_of_links = []
    with open(r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\pik_links.txt', 'r') as file:
        for item in file:
            list_of_links.append(item[:-1])

    return list_of_links


def operation(url, retry=5):
    global check_consist_pages
    try:
        options = webdriver.ChromeOptions()
        options.add_argument('--ignore-certificate-errors-spki-list')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument('--ignore-certificate-error')
        options.add_argument('--ignore-urlfetcher-cert-requests')

        driver = webdriver.Chrome(options=options)

        driver.get(url)
        time.sleep(40)

        pahe = driver.find_elements(By.CSS_SELECTOR, '.sc-gsnTZi.fWJuXR')

        count = len(pahe)

        while count != 0:
            count = 0

            for item in pahe:
                if ('Показать' in item.text):
                    actions = ActionChains(driver)
                    actions.move_to_element(item).perform()
                    time.sleep(2)
                    item.click()
                    time.sleep(2)
                    count += 1

            pahe = driver.find_elements(
                By.CSS_SELECTOR, '.sc-gsnTZi.fWJuXR')

        response = driver.page_source

        list_refer_storehouse_on_page = []
        soup = BeautifulSoup(response, 'lxml')
        for item in (soup.find_all('div', class_='sc-jVCNtP bVQpZB')):
            list_refer_storehouse_on_page.append(
                'https://www.pik.ru' + item.find('a').get('href'))

        driver.close()
        driver.quit()

        with open(fr'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\data_pages\{url.split('/')[-2]}.txt', 'w') as file:
            for link in list_refer_storehouse_on_page:
                file.write(f'{link}\n')

    except Exception as ex:
        time.sleep(5)
        if retry != 0:
            return options(url, retry=(retry-1))
        else:
            print(url, '  khyinya')
    print(f'{url.split('/')[-2]}.txt done')


def all_storegouse():
    global all_storehouse
    path = r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\data_pages'
    ways = os.listdir(path)

    all_storehouse = []
    for items in ways:
        with open(path+fr'\{items}', 'r') as file:
            for item in file:
                all_storehouse.append([items, item[:-1]])


def try_retry(index, item, retry=5):
    try:
        req = requests.get(item[1])
        soup = BeautifulSoup(req.text, 'lxml')
        try:
            order = soup.find('div', class_='sc-dHhosm bkzgjO').text
        except:
            order = 'Not'
        all_storehouse[index].append(order)

        req = requests.get(
            r'https://api.pik.ru/v2/flat/mortgage?flatId='+item[1][30:])
        js = json.loads(req.text)
        try:
            price = js['benefits']['cash'][0]['benefitPrice']
            all_storehouse[index].append(price)
        except:
            price = 'нет нашлось элемента'
            all_storehouse[index].append(price)
        try:
            price_m = js['benefits']['cash'][0]['priceMeter']
            all_storehouse[index].append(price_m)
        except:
            price_m = 'не нашлось элемента'
            all_storehouse[index].append(price_m)
    except Exception as ex:
        time.sleep(random.randint(5, 10))
        if retry != 0:
            try_retry(index, url, retry=(retry-1))
        else:
            print(f'{index} khinya')
    print(f'{index} done. ost {len(all_storehouse)-index}')


def get_info():
    global all_storehouse
    print(f'col-vo link {len(all_storehouse)}')
    for index, item in enumerate(all_storehouse):

        try_retry(index, item)


def write_excel():
    path = r'C:\Users\kraz1\OneDrive\Рабочий стол\Антон\python\парсинг\кладовочная\info.xlsx'

    if os.path.isfile(path) == True:
        os.remove(path)

    wb = ol.Workbook()
    wb.create_sheet('storehouse')
    wb.remove(wb['Sheet'])
    sheet = wb['storehouse']
    for index, item in enumerate(all_storehouse[0]):
        sheet.cell(row=1, column=1+index).value = 'lol.kek'
    for index_row, item_r in enumerate(all_storehouse):
        for index_column, item_c in enumerate(item_r):
            sheet.cell(row=2+index_row,  column=1 +
                       index_column).value = item_c
    wb.save(path)
    wb.close()


def main():
    print('1 часть н')
    list_links = get_list_objects_links(url)
    p = Pool(processes=5)
    p.map(operation, list_links)
    print('1 часть к')
    all_storegouse()
    get_info()
    write_excel()
    check_consist_pages_def()
    print('done')


if __name__ == '__main__':
    main()
